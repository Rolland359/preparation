# coding : "UTF-8"
import openpyxl
from openpyxl.styles import Alignment
import os

def extract_elements(sheet, start_row, col_indices):
    elements = []
    num = 1
    for i in range(start_row, sheet.max_row + 1):
        if sheet.cell(i, col_indices[-1]).value is not None:
            elements.append([num] + [(sheet.cell(i, col).value.strip().replace("\n","|") if sheet.cell(i, col).value else "") for col in col_indices])
            num += 1
    return elements

def extract_dossier(workbook):
    return str(workbook["Contrôle Global"]['D2'].value).zfill(4)

def ajuster_largeur_lignes(sheet):
    for row in sheet.iter_rows():
        max_height = 0
        for cell in row:
            if cell.value:
                lines = str(cell.value).count("\n")+1
                if lines > max_height:
                    max_height = lines
                cell.alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[row[0].row].height = max_height * 15

def remplissage(dossier_source, dossier, liste_elm_PM, liste_elm_BAT, fichier_gabarit):
    fichier_destination = f"{dossier_source}/ENR021_Checklist_RC_{dossier}.xlsx"
    wb_dest = openpyxl.load_workbook(fichier_gabarit, keep_vba=True)
    sheet_PM = wb_dest["Retour_clients_PM"]
    sheet_BAT = wb_dest["Retour_clients_2D-3D"]
    
    for row_PM, elm in enumerate(liste_elm_PM, start=3):
        sheet_PM.cell(row_PM, 1).value = elm[0]
        sheet_PM.cell(row_PM, 2).value = elm[1]
    
    for row_BAT, elm in enumerate(liste_elm_BAT, start=3):
        sheet_BAT.cell(row_BAT, 1).value = elm[0]
        sheet_BAT.cell(row_BAT, 2).value = elm[1]
        sheet_BAT.cell(row_BAT, 3).value = elm[2]
    
    ajuster_largeur_lignes(sheet_PM)
    ajuster_largeur_lignes(sheet_BAT)

    wb_dest.save(fichier_destination)
    wb_dest.close()

def main():
    dossier_source = input("Entrez le lien vers le fichier : ")
    fichier_source = next((os.path.join(dossier_source, elm) for elm in os.listdir(dossier_source) if "Grille contrôle Globale" in elm and "xls" in elm), None)
    
    if not fichier_source:
        print("Fichier source introuvable.")
        return

    wb = openpyxl.load_workbook(fichier_source)
    dossier = extract_dossier(wb)
    liste_elm_PM = extract_elements(wb["Contrôle PM"], 10, [7])
    liste_elm_BAT = extract_elements(wb["Contrôle Global"], 11, [2, 8])
    wb.close()

    remplissage(dossier_source, dossier, liste_elm_PM, liste_elm_BAT, fichier_gabarit)
    
    try:
        os.startfile(dossier_source)
    except Exception as e:
        print(f"Erreur survenue {e}")

fichier_gabarit = "C:/Futurmap/Outils/Preparation/Gestion de Patrimoine/RC/ENR021_Checklist_RC_XXXX.xlsx"

if __name__ == "__main__":
    os.system("title RC_GP")
    os.system("color a")
    main()
    while True:
        reponse = input("Voulez vous effectué un autre opération ? o/n :")
        if reponse.lower() == "o":
            main()
        else:
            print("Fin du programme ...")
            break