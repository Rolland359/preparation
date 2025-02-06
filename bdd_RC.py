# coding : "UTF-8"
import openpyxl
import sqlite3

def extract_elements(sheet, start_row, col_indices):
    return [
        [sheet.cell(i, col).value for col in col_indices]
        for i in range(start_row, sheet.max_row)
        if sheet.cell(i, col_indices[-1]).value
    ]

def extract_header(sheet,top_row,bottom_row, col_indices):
    elements = ["DOSSIER"]
    elements.extend(
        sheet.cell(i, col_indices).value
        for i in range(top_row, bottom_row)
        if sheet.cell(i, col_indices).value
    )
    return elements

def extract_dossier(workbook):
    return str(workbook["Contrôle Global"]['D2'].value).zfill(4)


nom_fichier = "Grille contrôle Globale 135.xlsx"
wb = openpyxl.load_workbook(nom_fichier)
sheet_BAT = wb["Contrôle Global"]
sheet_PM = wb["Contrôle PM"]

en_tete_bdd_BAT = extract_header(sheet_BAT,11,60,4)
en_tete_bdd_PM = extract_header(sheet_PM,)
print(en_tete_bdd_BAT)
print("\n")
print("-"*50)

dossier = extract_dossier(wb)
print(dossier)

dico_BAT = {"DOSSIER":dossier}
for row in range(11,sheet_BAT.max_row):
    cell_value = sheet_BAT.cell(row,8).value
    if cell_value:
        dico_BAT[sheet_BAT.cell(row,4).value] = cell_value

print("\n",dico_BAT)


wb.close()