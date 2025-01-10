# coding : UTF-8
import os
import subprocess
import openpyxl as xl

# 1. recuperer la FP : 
flies = os.listdir()
fiche_pat =str([fp for fp in flies if ".xlsx" in fp][0])
print(fiche_pat)

#charger la feuille de calcul
wb = xl.load_workbook(fiche_pat, read_only = False, data_only = True)
feuille_revet = wb.create_sheet('REVET', 1)
sheet = wb['ERP']


# matrice vierge pour le tableau des revêtements
Dict_tableau = {
    "A1":"DESIGNATION",   "B1":"SOL","C1":"PLAFOND","D1":"MUR",
    "A2":"PIECE SECHE",   "B2":"",   "C2":"",       "D2":"",
    "A3":"PIECE HUMIDE",  "B3":"",   "C3":"",       "D3":"",
    "A4":"PARTIE COMMUNE","B4":"",   "C4":"",       "D4":"",
    "A5":"PARKING",       "B5":"",   "C5":"",       "D5":"",
    "A6":"ANNEXE",        "B6":"",   "C6":"",       "D6":"",
}
# inserer les entêtes du tableau :
couleur_fond_cellule = "FF0000"
couleur_bordure = "000000"
type_bordure = xl.styles.Side(style='thin')
for k,v in Dict_tableau.items():
    cellules = feuille_revet[k]
    if k[1] == 1:
       cellules.value = v
       #couleur de remplissage :
       remplissage = xl.styles.Fill(fill_type="solid", bgColor=couleur_fond_cellule)
       cellules.fill = remplissage
       #bordure de cellule:
       bordure = xl.styles.Border(left=type_bordure,right=type_bordure,top=type_bordure,bottom=type_bordure)
       cellules.border = bordure
    else :
        cellules.value = v
        #bordure de cellule:
        bordure = xl.styles.Border(left=type_bordure,right=type_bordure,top=type_bordure,bottom=type_bordure)
        cellules.border = bordure


# enregistrer le fichier
wb.save(fiche_pat)

# verification:
cellule = feuille_revet.cell(1,1)
print(cellule.value)

# fermeture du fichier xl
wb.close()

