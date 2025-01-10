# coding : "UTF-8"
import openpyxl as xl
import os


def decorticage_clm():
    decorticage = open("DECORTICAGE.txt","w")
    lien_DB = "Q:/GP/2024/B24-0532(GP3D_CLM_STN_BC01)/1_Donnees_Brutes/BC02/"
    os.chdir(lien_DB)
    for dossier in os.listdir():
        lien_temp = f"{lien_DB}/{dossier}/1. PIECES GENERALES/1.1. Documents généraux/1.1.1. Fiche patrimoniale/"
        os.chdir(lien_temp)
        for fichier in os.listdir():
            if "FP" in fichier:
                fiche_pat = xl.load_workbook(fichier)
                feuille = fiche_pat['SOCIAL']
                ligne = 0

                for row in range(20, feuille.max_row,1):
                    if feuille.cell(row,2).value=="N° de Batiment :":
                        ligne = row
                    else:
                        pass
                
                for colonne in range(3, feuille.max_column,1):
                    valeur = feuille.cell(ligne,colonne).value
                    if valeur:
                        try:
                            decorticage.write(F"BC02\t{dossier}\tCLM BC02 {dossier} BATIMENT {valeur}_3D\tBATIMENT {valeur}\t{feuille.cell(ligne+5,colonne).value}\n")
                            
                        except:
                            print("erreur liée à l'écriture du fichier decorticage")
                print(f"{dossier} décirtiqué avec succès.")
                fiche_pat.close()
            else:
                pass
    decorticage.close()

decorticage_clm()