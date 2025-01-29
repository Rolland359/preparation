# coding: utf-8
import os
import openpyxl

def find_file(keyword):
    return next((elm for elm in os.listdir() if keyword in elm), None)

def load_workbook(file, sheet_name):
    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        return wb, sheet
    except Exception as e:
        print(f"Error loading workbook or sheet: {e}")
        return None, None

def search_results(sheet, dossier):
    results = []
    for i in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(i, 1).value
        if cell_value and dossier in str(cell_value):
            results.append((i, cell_value, sheet.cell(i, 6).value, sheet.cell(i, 3).value.replace("\n", " ")))
    return results

def get_user_choice(results):
    dict_result = {str(elm[0]): {"Dossier": elm[1], "AGENCE": elm[2], "SITE": elm[3]} for elm in results}
    for elm in results:
        print(f"ID: {elm[0]} \tDossier : {elm[1]} \tAGENCE : {elm[2]} \tSITE : {elm[3]}")
    id_choice = input("Entrer l'ID qui correspond exactemant à votre dossier : ")
    return dict_result[id_choice]

def copy_sheet_data(source_sheet, target_file):
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active
    new_worksheet.title = source_sheet.title

    for row in source_sheet.iter_rows(values_only=True):
        new_worksheet.append(row)

    new_workbook.save(target_file)

def normalize_string(s):
    return s.replace("-", "").replace("  ", " ").upper()

if __name__ == "__main__":
    os.system("title CCTP FINDER")
    os.system("color a")
    annexe_1 = find_file("ANNEXE 1")
    if not annexe_1:
        print("ANNEXE 1 file not found.")
        exit(1)

    dossier = input("Entrez le numéro du dossier : ")

    wb1, sheet1 = load_workbook(annexe_1, "PATRIMOINE COMPLET")
    liste_result = search_results(sheet1, dossier)
    wb1.close()

    if len(liste_result) > 1:
        print("Plusieurs Dossier découvert ...")
        choice = get_user_choice(liste_result)
        dossier_vrai, agence, adresse = choice["Dossier"], choice["AGENCE"], choice["SITE"].split("(")[0].strip()
    elif len(liste_result) == 1:
        dossier_vrai, agence, adresse = liste_result[0][1], liste_result[0][2], liste_result[0][3].split("(")[0].strip()
        print(liste_result[0][0], "Dossier : ", liste_result[0][1], "\tAGENCE : ", liste_result[0][2])
    else:
        print("Aucun Dossier trouvé.")
        exit(1)

    agence = agence.replace("CŒUR", "COEUR").replace("D'AUB", "AUB")
    adresse = adresse.replace("SAINT", "ST")

    print("AGENCE : ", agence)

    nom = find_file(agence)
    print(nom)
    if not nom:
        print(f"Aucun fichier annexe 2 trouvé pour l'agence : {agence}")
        exit(1)

    wb2 = openpyxl.load_workbook(nom)
    feuille_trouve = False

    normalized_adresse = normalize_string(adresse)
    for ws in wb2.sheetnames:
        ws_adresse = normalize_string(ws.split("HP")[0].strip())
        ws_dossier = ws.split("HP")
        if ws_adresse in normalized_adresse or normalized_adresse in ws_adresse and str(dossier_vrai) in ws_dossier:
            print("Feuille trouvé : ", ws)
            feuille_trouve = True
            feuille2 = wb2[ws]
            break

    if not feuille_trouve:
        print("Aucun Feuille trouvé ...")
    else:
        copy_sheet_data(feuille2, f'{ws}.xlsx')

    wb2.close()

    input()